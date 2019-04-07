from officenotes.models import OfficeNote
from order.models import Order
from customer.models import Customer
from graph.models import Workshop, DateRange

from random import randint
import openpyxl
import os
myhost = os.environ['COMPUTERNAME']

if myhost == "BOB":
    files_list = {
        'notes':'C:\\Users\\Yegor\\Dropbox\\ПДО_Производство\\Служебные записки.xlsx', 
        'graph':'C:\\Users\\Yegor\\Dropbox\\ПДО_Производство\\График план производства\\План производства.xlsx'
        }
elif myhost == "PDO-PRO":
    files_list = {
        'notes':'Z:\\Служебные записки.xlsx',
        'graph':'Z:\\График план производства\\План производства.xlsx'
        }

class Unit:
    def __init__(self, ready, tableid, shipment_from, shipment_to, product, counterparty, ordernum, quantity, firstofficenote, otherofficenote, date, datereceiving, pickingplan, pickingpercent, pickingfact, shippingplan, shippingpercent, shippingfact, engineeringplan, engineeringpercent, engineeringfact, drawingchangepercent, drawingchangefact, materialplan, materialfact):
        """Constructor for Unit"""
        if ready == '+':
            self.ready = True
        else:
            self.ready = False
        self.tableid = tableid # id в таблице 
        self.shipment_from = shipment_from # Отгрузка от 
        self.shipment_to = shipment_to # Отгрузка до 
        self.product = product # продукция 
        self.counterparty = counterparty # контрагент 
        self.ordernum = ordernum # № Заказа 
        self.quantity = quantity # Кол-во 
        self.firstofficenote = firstofficenote # № СЗ 
        self.otherofficenote = otherofficenote # № СЗ с изменениями 
        self.date = date # Дата СЗ 
        self.datereceiving = datereceiving # Дата СЗ Факт 
        self.pickingplan = pickingplan # Комплектовочные План 
        self.pickingpercent = pickingpercent # Комплектовочные % 
        self.pickingfact = pickingfact # Комплектовочные Факт 
        self.shippingplan = shippingplan # Отгрузочные План 
        self.shippingpercent = shippingpercent # Отгрузочные % 
        self.shippingfact = shippingfact # Отгрузочные Факт 
        self.engineeringplan = engineeringplan # Конструкторская План 
        self.engineeringpercent = engineeringpercent # Конструкторская % 
        self.engineeringfact = engineeringfact # Конструкторская Факт 
        self.drawingchangepercent = drawingchangepercent # Изменения чертежей % 
        self.drawingchangefact = drawingchangefact # Изменения чертежей Факт 
        self.materialplan = materialplan # Материалы План 
        self.materialfact = materialfact # Материалы Факт 

class OfficeNotePath:
    def __init__(self, officenote, fpath):
        """Constructor for OfficeNotePath"""
        self.officenote = officenote # № СЗ
        self.fpath = fpath # № СЗ

class CWorkshop:
    def __init__(self, numname, name, fullname):
        """Constructor for CWorkshop"""
        self.numname = numname # Номер цеха
        self.name = name # Название 
        self.fullname = fullname # Полное название

class Graph:
    def __init__(self, dorder, twodatestart, twodateend, onedatestart, onedateend, sevendatestart, sevendateend, fourdatestart, fourdateend):
        """Constructor for OfficeNotePath"""
        self.dorder = dorder # № Заказа
        self.twodatestart = twodatestart # 
        self.twodateend = twodateend # 
        self.onedatestart = onedatestart # 
        self.onedateend = onedateend # 
        self.sevendatestart = sevendatestart # 
        self.sevendateend = sevendateend # 
        self.fourdatestart = fourdatestart # 
        self.fourdateend = fourdateend # 

def append_graph(fpath):
    yes_processed_rows = 0
    not_processed_rows = 0
    log_text = ''

    try:
        work_wb = openpyxl.load_workbook(filename=fpath)
    except:
        print('trabl')
        exit(0)
    sheet = work_wb['Даты']
    nunits = []
    for row in range(2, sheet.max_row+1):
        unit = Graph(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            sheet.cell(row=row, column=3).value,
            sheet.cell(row=row, column=4).value,
            sheet.cell(row=row, column=5).value,
            sheet.cell(row=row, column=6).value,
            sheet.cell(row=row, column=7).value,
            sheet.cell(row=row, column=8).value,
            sheet.cell(row=row, column=9).value
        )
        nunits.append(unit)
    work_wb.close()

    yes_processed_rows += len(nunits)
    not_processed_rows += 0
    log_text += 'Прочитано %s строк на листе "Даты" в файле: %s \n' % (len(nunits), fpath)

    sa = []
    shop = Workshop.objects.all()
    for un in nunits:
        drorder = Order.objects.filter(ordernum=un.dorder)
        for order in drorder:
            for sh in shop:
                if sh.numname == 102:
                    if un.twodatestart != None:
                        sa.append(DateRange(
                            workshop = sh,
                            order = order,
                            datestart = un.twodatestart,
                            dateend = un.twodateend
                        ))
                if sh.numname == 101:
                    if un.onedatestart != None:
                        sa.append(DateRange(
                            workshop = sh,
                            order = order,
                            datestart = un.onedatestart,
                            dateend = un.onedateend
                        ))
                if sh.numname == 107:
                    if un.sevendatestart != None:
                        sa.append(DateRange(
                            workshop = sh,
                            order = order,
                            datestart = un.sevendatestart,
                            dateend = un.sevendateend
                        ))
                if sh.numname == 104:
                    if un.fourdatestart != None:
                        sa.append(DateRange(
                            workshop = sh,
                            order = order,
                            datestart = un.fourdatestart,
                            dateend = un.fourdateend
                        ))
    DateRange.objects.bulk_create(sa)
    return [yes_processed_rows, not_processed_rows, log_text]

def append_notes(fpath):
    yes_processed_rows = 0
    not_processed_rows = 0
    log_text = ''

    try:
        work_wb = openpyxl.load_workbook(filename=fpath)
    except:
        print('trabl')
        exit(0)
    sheet = work_wb['Просчет']
    nunits = []
    for row in range(2, sheet.max_row+1):
        unit = Unit(
            sheet.cell(row=row, column=3).value,
            sheet.cell(row=row, column=2).value,
            sheet.cell(row=row, column=4).value,
            sheet.cell(row=row, column=5).value,
            sheet.cell(row=row, column=8).value,
            sheet.cell(row=row, column=9).value,
            sheet.cell(row=row, column=10).value,
            sheet.cell(row=row, column=11).value,
            sheet.cell(row=row, column=12).value,
            sheet.cell(row=row, column=13).value,
            sheet.cell(row=row, column=15).value,
            sheet.cell(row=row, column=16).value,
            sheet.cell(row=row, column=20).value,
            sheet.cell(row=row, column=21).value,
            sheet.cell(row=row, column=22).value,
            sheet.cell(row=row, column=26).value,
            sheet.cell(row=row, column=27).value,
            sheet.cell(row=row, column=28).value,
            sheet.cell(row=row, column=32).value,
            sheet.cell(row=row, column=33).value,
            sheet.cell(row=row, column=34).value,
            sheet.cell(row=row, column=36).value,
            sheet.cell(row=row, column=37).value,
            sheet.cell(row=row, column=38).value,
            sheet.cell(row=row, column=39).value,
            )
        nunits.append(unit)

    yes_processed_rows += len(nunits)
    not_processed_rows += 0
    log_text += 'Прочитано %s строк на листе "Просчет" в файле: %s \n' % (len(nunits), fpath)

    pathsheet = work_wb['Файлы']
    officepath = []
    for row in range(2, pathsheet.max_row+1):
        unitfp = OfficeNotePath(
            pathsheet.cell(row=row, column=2).value,
            pathsheet.cell(row=row, column=4).value,
        )
        officepath.append(unitfp)

    yes_processed_rows += len(officepath)
    not_processed_rows += 0
    log_text += 'Прочитано %s строк на листе "Файлы" в файле: %s \n' % (len(officepath), fpath)

    shopsheet = work_wb['Цех']
    shop = []
    for row in range(2, shopsheet.max_row+1):
        print(row)
        unitfp = CWorkshop(
            shopsheet.cell(row=row, column=1).value,
            shopsheet.cell(row=row, column=2).value,
            shopsheet.cell(row=row, column=3).value
        )
        shop.append(unitfp)

    yes_processed_rows += len(shop)
    not_processed_rows += 0
    log_text += 'Прочитано %s строк на листе "Цех" в файле: %s \n' % (len(shop), fpath)

    work_wb.close()

    objs = [
    Order(
        ready = nunit.ready,
        product = nunit.product,
        tableid = nunit.tableid,
        shipmentfrom = nunit.shipment_from,
        shipmentto = nunit.shipment_to,
        ordernum = nunit.ordernum,
        quantity = nunit.quantity,
        pickingplan = nunit.pickingplan,
        pickingpercent = nunit.pickingpercent,
        pickingfact = nunit.pickingfact,
        shippingplan = nunit.shippingplan,
        shippingpercent = nunit.shippingpercent,
        shippingfact = nunit.shippingfact,
        engineeringplan = nunit.engineeringplan,
        engineeringpercent = nunit.engineeringpercent,
        engineeringfact = nunit.engineeringfact,
        drawingchangepercent = nunit.drawingchangepercent,
        drawingchangefact = nunit.drawingchangefact,
        materialplan = nunit.materialplan,
        materialfact = nunit.materialfact,
        firstofficenote = OfficeNote.objects.get_or_create(
                num = nunit.firstofficenote,
                date = nunit.date,
                datereceiving = nunit.datereceiving,
                oncustomer = Customer.objects.get_or_create(
                    name = nunit.counterparty
                    )[0]
                )[0]
        )
    for nunit in nunits
    ]
    Order.objects.bulk_create(objs)

    notes = Order.objects.all()
    
    for nunit in nunits:
        if nunit.otherofficenote != None:
            note = notes.get(tableid=nunit.tableid)
            for othernum in nunit.otherofficenote.split(','):
                note.otherofficenote.add(
                    OfficeNote.objects.get_or_create(num = othernum)[0]
                )

    ubdatesofficenotes = OfficeNote.objects.all()
    sa = []
    for ufpath in officepath:
        note = ubdatesofficenotes.get(num=ufpath.officenote)
        note.filepath = ufpath.fpath
        sa.append(note)
    OfficeNote.objects.bulk_update(sa, ['filepath',])

    objs = [
    Workshop(
        numname = nunit.numname,
        name = nunit.name,
        fullname = nunit.fullname,
        )
    for nunit in shop
    ]
    Workshop.objects.bulk_create(objs)

    return [yes_processed_rows, not_processed_rows, log_text]

def append_base(flist):
    yes_processed_rows = 0
    not_processed_rows = 0
    log_text = ''

    status = []
    for name_model, fpath in flist.items():
        print(fpath)
        if name_model == 'notes':
            print('notes')
            status = append_notes(fpath)
            yes_processed_rows += status[0]
            not_processed_rows += status[1]
            log_text += status[2]

        if name_model == 'graph':
            print('graph')
            status = append_graph(fpath)
            yes_processed_rows += status[0]
            not_processed_rows += status[1]
            log_text += status[2]
    return [yes_processed_rows, not_processed_rows, log_text]

def delete_base():
    yes_processed_rows = 0
    not_processed_rows = 0
    log_text = ''

    custom_deleted = Customer.objects.all().delete()
    yes_processed_rows += 0
    not_processed_rows += custom_deleted[0]
    log_text += 'Удалено %s записей из "Customer" \n' % custom_deleted[0]

    workshop_deleted = Workshop.objects.all().delete()
    yes_processed_rows += 0
    not_processed_rows += workshop_deleted[0]
    log_text += 'Удалено %s записей из "Workshop" \n' % workshop_deleted[0]

    officenote_deleted = OfficeNote.objects.all().delete()
    yes_processed_rows += 0
    not_processed_rows += officenote_deleted[0]
    log_text += 'Удалено %s записей из "OfficeNote" \n' % officenote_deleted[0]

    return [yes_processed_rows, not_processed_rows, log_text]

def update():
    yes_processed_rows = 0
    not_processed_rows = 0
    log_text = ''

    deletebd = delete_base()
    yes_processed_rows += deletebd[0]
    not_processed_rows += deletebd[1]
    log_text += deletebd[2]

    appendbd = append_base(files_list)
    yes_processed_rows += appendbd[0]
    not_processed_rows += appendbd[1]
    log_text += appendbd[2]
    return yes_processed_rows, not_processed_rows, log_text